import json
import sys
import time
import argparse
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, List, Tuple, Optional

import urllib.parse
import urllib.request


ROOT = Path(__file__).resolve().parents[1]
SECRETS_PATH = ROOT / "secrets.local.json"
DOCS_DIR = ROOT / "docs"
REPORTS_DIR = DOCS_DIR / "reports"


def http_get(url: str) -> Dict[str, Any]:
    try:
        with urllib.request.urlopen(url) as r:  # nosec - controlled domains (graph.facebook.com)
            data = r.read()
            return json.loads(data.decode("utf-8"))
    except urllib.error.HTTPError as e:  # type: ignore[attr-defined]
        try:
            body = e.read().decode("utf-8")  # type: ignore[attr-defined]
            return json.loads(body)
        except Exception:
            raise


def load_secrets() -> Dict[str, Any]:
    if not SECRETS_PATH.exists():
        raise FileNotFoundError(f"Missing secrets file: {SECRETS_PATH}")
    return json.loads(SECRETS_PATH.read_text(encoding="utf-8"))


def paginate(url: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    next_url: Optional[str] = url
    while next_url:
        obj = http_get(next_url)
        data = obj.get("data", [])
        out.extend(data)
        paging = obj.get("paging", {})
        next_url = paging.get("next")
        # be gentle
        time.sleep(0.15)
    return out


def extract_shortcode(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    try:
        u = url.split("?")[0].rstrip("/")
        parts = u.split("/")
        # expect .../reel/{code} or /p/{code} or /tv/{code}
        if len(parts) >= 2:
            code = parts[-1]
            if code:
                return code
    except Exception:
        return None
    return None

def ig_media_list(ig_user_id: str, page_token: str, days: int = 90) -> List[Dict[str, Any]]:
    since = int((datetime.utcnow() - timedelta(days=days)).timestamp())
    fields = (
        "id,caption,media_type,media_product_type,permalink,timestamp,like_count,comments_count"
    )
    base = f"https://graph.facebook.com/v22.0/{ig_user_id}/media?fields={urllib.parse.quote(fields)}&access_token={urllib.parse.quote(page_token)}&limit=50"
    media = paginate(base)
    # filter by date
    out = []
    for m in media:
        ts = m.get("timestamp")
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00")) if ts else None
        except Exception:
            dt = None
        if dt and dt.timestamp() >= since:
            out.append(m)
    return out


MEDIA_METRICS = [
    "reach",
    "likes",
    "comments",
    "saved",
    "shares",
    "views",
    "total_interactions",
    # try follows where supported; ignore failures
]


def ig_media_insights(media_id: str, page_token: str) -> Dict[str, Any]:
    metrics = ",".join(MEDIA_METRICS)
    url = (
        f"https://graph.facebook.com/v22.0/{media_id}/insights?metric="
        f"{urllib.parse.quote(metrics)}&access_token={urllib.parse.quote(page_token)}"
    )
    try:
        obj = http_get(url)
        res: Dict[str, Any] = {}
        for item in obj.get("data", []):
            name = item.get("name")
            values = item.get("values", [])
            if values:
                res[name] = values[0].get("value")
        return res
    except Exception:
        return {}


def list_ad_accounts(user_token: str) -> List[Dict[str, Any]]:
    url = (
        "https://graph.facebook.com/v19.0/me/adaccounts?fields=id,account_id,name,currency"  # noqa: E501
        f"&access_token={urllib.parse.quote(user_token)}"
    )
    return paginate(url)


def ad_insights_by_ad_range(ad_account_id: str, user_token: str, since: str, until: str) -> List[Dict[str, Any]]:
    fields = (
        "ad_id,ad_name,campaign_name,adset_name,spend,impressions,reach,clicks,ctr,cpc,cpm,video_3_sec_views,date_start,date_stop,actions"
    )
    time_range = json.dumps({"since": since, "until": until})
    q = urllib.parse.urlencode(
        {
            "level": "ad",
            "time_range": time_range,
            "fields": fields,
            "limit": 200,
            "access_token": user_token,
        }
    )
    url = f"https://graph.facebook.com/v19.0/{ad_account_id}/insights?{q}"
    return paginate(url)


def ad_creative_info(ad_id: str, user_token: str) -> Dict[str, Any]:
    fields = urllib.parse.quote("creative{effective_object_story_id,instagram_permalink_url}")
    url = f"https://graph.facebook.com/v19.0/{ad_id}?fields={fields}&access_token={urllib.parse.quote(user_token)}"
    try:
        return http_get(url)
    except Exception:
        return {}


def fetch_ad_creatives_map(ad_account_id: str, user_token: str) -> Dict[str, str]:
    """Return map of ad_id -> instagram shortcode (derived from permalink), when available."""
    url = (
        f"https://graph.facebook.com/v19.0/{ad_account_id}/ads?"
        + urllib.parse.urlencode(
            {
                "fields": "id,creative{instagram_permalink_url,effective_object_story_id,object_story_id}",
                "limit": 200,
                "access_token": user_token,
            }
        )
    )
    data = paginate(url)
    out: Dict[str, str] = {}
    for a in data:
        ad_id = a.get("id")
        creative = a.get("creative") or {}
        permalink = creative.get("instagram_permalink_url")
        sc = extract_shortcode(permalink)
        if ad_id and sc:
            out[ad_id] = sc
    return out


def build_mapping_paid_to_media(ad_insights: List[Dict[str, Any]], user_token: str, *, ad_creatives: Optional[Dict[str, str]] = None, audit: Optional[List[Dict[str, Any]]] = None, account_id: Optional[str] = None) -> Dict[str, Dict[str, Any]]:
    mapping: Dict[str, Dict[str, Any]] = {}
    for row in ad_insights:
        ad_id = row.get("ad_id") or row.get("adid") or row.get("ad")
        if not ad_id:
            continue
        shortcode = None
        if ad_creatives is not None:
            shortcode = ad_creatives.get(str(ad_id))
        if not shortcode:
            info = ad_creative_info(ad_id, user_token)
            creative = info.get("creative", {}) if isinstance(info, dict) else {}
            permalink = creative.get("instagram_permalink_url")
            shortcode = extract_shortcode(permalink)
        if not shortcode:
            continue
        m = mapping.setdefault(shortcode, {
            "paid_spend": 0.0,
            "paid_impressions": 0,
            "paid_reach": 0,
            "paid_clicks": 0,
            "paid_video_views": 0,
            "paid_post_engagement": 0,
            "paid_video_3s": 0,
            "paid_profile_visits": 0,
            "ad_start": None,
            "ad_end": None,
        })
        # Aggregate
        try:
            m["paid_spend"] += float(row.get("spend", 0) or 0)
        except Exception:
            pass
        try:
            m["paid_impressions"] += int(row.get("impressions", 0) or 0)
        except Exception:
            pass
        try:
            m["paid_reach"] += int(row.get("reach", 0) or 0)
        except Exception:
            pass
        try:
            m["paid_clicks"] += int(row.get("clicks", 0) or 0)
        except Exception:
            pass
        try:
            m["paid_video_3s"] += int(row.get("video_3_sec_views", 0) or 0)
        except Exception:
            pass
        # actions array may include video_view, post_engagement, etc.
        for act in row.get("actions", []) or []:
            if act.get("action_type") == "video_view":
                try:
                    m["paid_video_views"] += int(act.get("value", 0) or 0)
                except Exception:
                    pass
            if act.get("action_type") in {"post_engagement", "post_interaction_gross"}:
                try:
                    m["paid_post_engagement"] += int(act.get("value", 0) or 0)
                except Exception:
                    pass
            if "profile" in str(act.get("action_type", "")) and "visit" in str(act.get("action_type", "")):
                try:
                    m["paid_profile_visits"] += int(act.get("value", 0) or 0)
                except Exception:
                    pass
        # duration window across ads referencing this post
        s = row.get("date_start")
        e = row.get("date_stop")
        if s and (m["ad_start"] is None or s < m["ad_start"]):
            m["ad_start"] = s
        if e and (m["ad_end"] is None or e > m["ad_end"]):
            m["ad_end"] = e
        time.sleep(0.05)
        if audit is not None:
            audit.append({
                "ad_account": account_id,
                "ad_id": ad_id,
                "permalink": permalink,
                "shortcode": shortcode,
                "campaign_name": row.get("campaign_name"),
                "adset_name": row.get("adset_name"),
                "date_start": s,
                "date_stop": e,
                "spend": row.get("spend"),
                "reach": row.get("reach"),
                "impressions": row.get("impressions"),
                "clicks": row.get("clicks"),
                "video_3s": row.get("video_3_sec_views"),
            })
    return mapping


def ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return
    except Exception:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])  # noqa: S603,S607


def write_excel(rows: List[Dict[str, Any]], out_path: Path, *, ig_trends: Optional[List[Dict[str, Any]]] = None, fb_posts: Optional[List[Dict[str, Any]]] = None, rows_posts: Optional[List[Dict[str, Any]]] = None, ad_audit: Optional[List[Dict[str, Any]]] = None):
    ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "רילס"
    # Right-to-left view
    ws.sheet_view.rightToLeft = True
    # Headers (Hebrew)
    headers = [
        "תאריך",
        "סוג תוכן",
        "כותרת",
        "קישור",
        "שיתופים",
        "שימורים",
        "תגובות",
        "לייקים",
        "צפיות (סה""כ)",
        "טווח הגעה (סה""כ)",
        "אינטראקציות (סה""כ)",
        "טווח הגעה ממומן",
        "חשיפות ממומנות",
        "קליקים ממומנים",
        "צפיות 3ש ממומן",
        "צפיות ממומן",
        "סכום ממומן (ILS)",
        "ביקורי פרופיל ממומן",
        "ימי קמפיין",
        "אינטראקציות אורגניות (הערכה)",
    ]
    ws.append(headers)

    align_right = Alignment(horizontal="right")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)

    for r in rows:
        ws.append([
            r.get("date"),
            r.get("media_type"),
            r.get("title"),
            r.get("permalink"),
            r.get("shares"),
            r.get("saved"),
            r.get("comments"),
            r.get("likes"),
            r.get("views"),
            r.get("reach"),
            r.get("total_interactions"),
            r.get("paid_reach"),
            r.get("paid_impressions"),
            r.get("paid_clicks"),
            r.get("paid_video_3s"),
            r.get("paid_video_views"),
            r.get("paid_spend"),
            r.get("paid_profile_visits"),
            r.get("ad_duration_days"),
            r.get("organic_interactions_est"),
        ])

    # Header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_right

    # Column alignment and width
    for col in ws.columns:
        for cell in col:
            cell.alignment = align_right
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 60)

    # Freeze header and add filterable table
    ws.freeze_panes = "A2"
    last_row = ws.max_row
    last_col = ws.max_column
    if last_row > 1:
        ref = f"A1:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName="IGReport", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

    # Sheet 2: IG posts (non-reels)
    if rows_posts:
        wsP = wb.create_sheet("פוסטים אינסטגרם")
        wsP.sheet_view.rightToLeft = True
        wsP.append(headers)
        for r in rows_posts:
            wsP.append([
                r.get("date"), r.get("media_type"), r.get("title"), r.get("permalink"),
                r.get("shares"), r.get("saved"), r.get("comments"), r.get("likes"),
                r.get("views"), r.get("reach"), r.get("total_interactions"),
                r.get("paid_impressions"), r.get("paid_clicks"), r.get("paid_video_views"),
                r.get("paid_spend"), r.get("organic_interactions_est"),
            ])
        for cell in wsP[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = align_right
        for col in wsP.columns:
            for cell in col: cell.alignment = align_right
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            wsP.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 60)
        wsP.freeze_panes = "A2"
        last_row = wsP.max_row; last_col = wsP.max_column
        if last_row > 1:
            ref = f"A1:{get_column_letter(last_col)}{last_row}"
            tP = Table(displayName="IGPosts", ref=ref)
            tP.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            wsP.add_table(tP)

    # Sheet 3: IG account trends (optional)
    if ig_trends:
        ws2 = wb.create_sheet("מגמות חשבון")
        ws2.sheet_view.rightToLeft = True
        headers2 = ["תאריך", "הגעה יומית", "ביקורי פרופיל (סה""כ)", "חשבונות מעורבים (סה""כ)", "אינטראקציות (סה""כ)", "עוקבים (סה""כ)"]
        ws2.append(headers2)
        for row in ig_trends:
            ws2.append([
                row.get("date"),
                row.get("reach"),
                row.get("profile_views"),
                row.get("accounts_engaged"),
                row.get("total_interactions"),
                row.get("follower_count"),
            ])
        for cell in ws2[1]:
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")
        for col in ws2.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal="right")
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 50)
        ws2.freeze_panes = "A2"
        last_row = ws2.max_row
        last_col = ws2.max_column
        if last_row > 1:
            ref = f"A1:{get_column_letter(last_col)}{last_row}"
            t2 = Table(displayName="IGTrends", ref=ref)
            t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws2.add_table(t2)

    # Sheet 4: Facebook posts (optional)
    if fb_posts:
        ws3 = wb.create_sheet("פייסבוק")
        ws3.sheet_view.rightToLeft = True
        headers3 = ["תאריך", "כותרת", "קישור", "חשיפות", "הגעה ייחודית", "משתמשים מעורבים", "קליקים"]
        ws3.append(headers3)
        for p in fb_posts:
            ws3.append([
                p.get("date"),
                p.get("title"),
                p.get("permalink_url"),
                p.get("post_impressions"),
                p.get("post_impressions_unique"),
                p.get("post_engaged_users"),
                p.get("post_clicks"),
            ])
        for cell in ws3[1]:
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right")
        for col in ws3.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal="right")
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws3.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 60)
        ws3.freeze_panes = "A2"
        last_row = ws3.max_row
        last_col = ws3.max_column
        if last_row > 1:
            ref = f"A1:{get_column_letter(last_col)}{last_row}"
            t3 = Table(displayName="FBPosts", ref=ref)
            t3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws3.add_table(t3)

    # Sheet 5: Ad audit (optional)
    if ad_audit:
        wsA = wb.create_sheet("מיפוי מודעות")
        wsA.sheet_view.rightToLeft = True
        hdrA = ["חשבון מודעות","מזהה מודעה","קישור אינסטגרם","קמפיין","סט מודעות","התחלה","סיום","סכום (ILS)","הגעה","חשיפות","קליקים","3ש צפיות"]
        wsA.append(hdrA)
        for a in ad_audit:
            wsA.append([
                a.get("ad_account"), a.get("ad_id"), a.get("permalink"), a.get("campaign_name"), a.get("adset_name"), a.get("date_start"), a.get("date_stop"), a.get("spend"), a.get("reach"), a.get("impressions"), a.get("clicks"), a.get("video_3s")
            ])
        for cell in wsA[1]:
            cell.fill = PatternFill("solid", fgColor="D9E1F2"); cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="right")
        for col in wsA.columns:
            for cell in col: cell.alignment = Alignment(horizontal="right")
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            wsA.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 60)
        wsA.freeze_panes = "A2"
        last_row = wsA.max_row; last_col = wsA.max_column
        if last_row > 1:
            ref = f"A1:{get_column_letter(last_col)}{last_row}"
            tA = Table(displayName="AdAudit", ref=ref)
            tA.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            wsA.add_table(tA)

    # Save
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def ig_user_daily_series(ig_user_id: str, page_token: str, days: int = 30) -> List[Dict[str, Any]]:
    # Build daily series for reach and totals
    def fetch(metric: str, *, total: bool = False) -> List[Tuple[str, int]]:
        since = int((datetime.utcnow() - timedelta(days=days)).timestamp())
        until = int(datetime.utcnow().timestamp())
        params = [
            ("metric", metric),
            ("period", "day"),
            ("since", str(since)),
            ("until", str(until)),
            ("access_token", page_token),
        ]
        if total:
            params.append(("metric_type", "total_value"))
        q = urllib.parse.urlencode(params)
        url = f"https://graph.facebook.com/v22.0/{ig_user_id}/insights?{q}"
        obj = http_get(url)
        seq: List[Tuple[str, int]] = []
        try:
            data = obj.get("data", [])[0].get("values", [])
            for v in data:
                seq.append((v.get("end_time", "")[:10], int(v.get("value", 0) or 0)))
        except Exception:
            pass
        return seq

    # metrics
    reach = fetch("reach", total=False)
    profile = fetch("profile_views", total=True)
    engaged = fetch("accounts_engaged", total=True)
    interactions = fetch("total_interactions", total=True)
    followers = fetch("follower_count", total=True)

    # Merge by date
    dates = {d for d, _ in reach} | {d for d, _ in profile} | {d for d, _ in engaged} | {d for d, _ in interactions} | {d for d, _ in followers}
    out: List[Dict[str, Any]] = []
    for d in sorted(dates):
        row = {"date": d}
        row["reach"] = next((v for dd, v in reach if dd == d), None)
        row["profile_views"] = next((v for dd, v in profile if dd == d), None)
        row["accounts_engaged"] = next((v for dd, v in engaged if dd == d), None)
        row["total_interactions"] = next((v for dd, v in interactions if dd == d), None)
        row["follower_count"] = next((v for dd, v in followers if dd == d), None)
        out.append(row)
    return out

def fb_page_posts_with_insights(page_id: str, page_token: str, days: int = 30) -> List[Dict[str, Any]]:
    since_dt = datetime.now(timezone.utc) - timedelta(days=days)
    fields = "id,created_time,permalink_url,message"
    base = f"https://graph.facebook.com/v19.0/{page_id}/posts?fields={urllib.parse.quote(fields)}&access_token={urllib.parse.quote(page_token)}&limit=50"
    posts = paginate(base)
    out: List[Dict[str, Any]] = []
    for p in posts:
        ct = p.get("created_time")
        try:
            dt = datetime.fromisoformat(ct.replace("Z", "+00:00")) if ct else None
        except Exception:
            dt = None
        if not dt or dt < since_dt:
            continue
        pid = p.get("id")
        metrics = "post_impressions,post_impressions_unique,post_engaged_users,post_clicks"
        url = f"https://graph.facebook.com/v19.0/{pid}/insights?metric={urllib.parse.quote(metrics)}&access_token={urllib.parse.quote(page_token)}"
        ins = http_get(url)
        row = {
            "date": ct[:10] if ct else "",
            "title": (p.get("message") or "").splitlines()[0][:60] if p.get("message") else "",
            "permalink_url": p.get("permalink_url"),
        }
        for item in ins.get("data", []) or []:
            name = item.get("name")
            vals = item.get("values", [])
            if vals:
                row[name] = vals[0].get("value")
        out.append(row)
        time.sleep(0.05)
    return out


def main():
    ap = argparse.ArgumentParser(description="Generate IG + FB performance report")
    ap.add_argument("--ig-days", type=int, default=365, help="Days of IG media to include (default 365)")
    ap.add_argument("--ads-days", type=int, default=365, help="Days of ad insights to include (default 365)")
    ap.add_argument("--outfile", type=str, default="", help="Optional output filename")
    args = ap.parse_args()

    secrets = load_secrets()
    page_token = secrets.get("instagram", {}).get("page_access_token") or secrets.get("facebook_page", {}).get("access_token")
    ig_user_id = secrets.get("instagram", {}).get("user_id")
    user_token = secrets.get("facebook_user", {}).get("access_token")
    if not page_token or not ig_user_id or not user_token:
        print("Missing required tokens in secrets.local.json (instagram.page_access_token, instagram.user_id, facebook_user.access_token)", file=sys.stderr)
        return 2

    # IG media and insights
    media = ig_media_list(ig_user_id, page_token, days=int(args.ig_days))
    enriched: List[Dict[str, Any]] = []
    enriched_posts: List[Dict[str, Any]] = []
    for m in media:
        ins = ig_media_insights(m["id"], page_token)
        cap = (m.get("caption") or "").strip()
        # title: first 50 chars
        title = cap.splitlines()[0][:60] if cap else m.get("permalink")
        row = {
            "id": m.get("id"),
            "date": (m.get("timestamp") or "")[:10],
            "media_type": m.get("media_type"),
            "media_product_type": m.get("media_product_type"),
            "title": title,
            "permalink": m.get("permalink"),
            "likes": ins.get("likes", m.get("like_count")),
            "comments": ins.get("comments", m.get("comments_count")),
            "saved": ins.get("saved", 0),
            "shares": ins.get("shares", 0),
            "views": ins.get("views", 0),
            "reach": ins.get("reach", 0),
            "total_interactions": ins.get("total_interactions", 0),
        }
        # Split into reels vs posts
        if (m.get("media_product_type") or "").upper() == "REELS":
            enriched.append(row)
        else:
            enriched_posts.append(row)
        time.sleep(0.05)

    # Paid mapping via Marketing API
    adacts = list_ad_accounts(user_token)
    paid_map: Dict[str, Dict[str, Any]] = {}
    ad_audit: List[Dict[str, Any]] = []
    if adacts:
        since = (datetime.now().date() - timedelta(days=int(args.ads_days))).strftime("%Y-%m-%d")
        until = datetime.now().date().strftime("%Y-%m-%d")
        for acct in adacts:
            act_id = acct.get("id")
            if not act_id:
                continue
            creatives_map = fetch_ad_creatives_map(act_id, user_token)
            ads_rows = ad_insights_by_ad_range(act_id, user_token, since, until)
            # accumulate into the paid_map
            part = build_mapping_paid_to_media(ads_rows, user_token, ad_creatives=creatives_map, audit=ad_audit, account_id=act_id)
            for k, v in part.items():
                agg = paid_map.setdefault(k, {})
                for fk in v:
                    if fk in {"ad_start","ad_end"}:
                        # keep min/max
                        if fk == "ad_start":
                            agg[fk] = min(filter(None, [agg.get(fk), v.get(fk)])) if agg.get(fk) else v.get(fk)
                        else:
                            agg[fk] = max(filter(None, [agg.get(fk), v.get(fk)])) if agg.get(fk) else v.get(fk)
                        continue
                    try:
                        agg[fk] = (agg.get(fk, 0) or 0) + (v.get(fk, 0) or 0)
                    except Exception:
                        agg[fk] = v.get(fk)

    # Merge paid into media rows using permalink key
    for r in enriched:
        sc = extract_shortcode(r.get("permalink"))
        p = paid_map.get(sc or "", {})
        r.update({
            "paid_impressions": p.get("paid_impressions", 0),
            "paid_reach": p.get("paid_reach", 0),
            "paid_clicks": p.get("paid_clicks", 0),
            "paid_video_3s": p.get("paid_video_3s", 0),
            "paid_video_views": p.get("paid_video_views", 0),
            "paid_spend": round(float(p.get("paid_spend", 0.0)), 2),
            "paid_post_engagement": p.get("paid_post_engagement", 0),
            "paid_profile_visits": p.get("paid_profile_visits", 0),
        })
        try:
            ti = int(r.get("total_interactions") or 0)
            pe = int(p.get("paid_post_engagement") or 0)
            r["organic_interactions_est"] = max(0, ti - pe)
        except Exception:
            r["organic_interactions_est"] = None
        # duration days
        try:
            if p.get("ad_start") and p.get("ad_end"):
                s = datetime.strptime(p["ad_start"], "%Y-%m-%d")
                e = datetime.strptime(p["ad_end"], "%Y-%m-%d")
                r["ad_duration_days"] = (e - s).days + 1
            else:
                r["ad_duration_days"] = None
        except Exception:
            r["ad_duration_days"] = None

    # IG account trends (last 30d)
    ig_trends = ig_user_daily_series(ig_user_id, page_token, days=30)

    # Facebook page posts (last 30d)
    page_id = secrets.get("facebook_page", {}).get("page_id") or secrets.get("instagram", {}).get("page_id")
    fb_posts = fb_page_posts_with_insights(page_id, page_token, days=30) if page_id else []

    # Write Excel
    ts = datetime.now().strftime("%Y%m%d")
    if args.outfile:
        out_path = Path(args.outfile)
        out_path = out_path if out_path.is_absolute() else (REPORTS_DIR / out_path)
    else:
        out_path = REPORTS_DIR / f"IG_Report_{ts}.xlsx"
        if out_path.exists():
            out_path = REPORTS_DIR / f"IG_Report_{ts}_{datetime.now().strftime('%H%M%S')}.xlsx"
    write_excel(enriched, out_path, ig_trends=ig_trends, fb_posts=fb_posts, rows_posts=enriched_posts, ad_audit=ad_audit)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
